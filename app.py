from flask import Flask, request, jsonify, send_from_directory
from pathlib import Path
from datetime import datetime
import sqlite3, csv
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / 'data'
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = DATA_DIR / 'app.db'
CSV_PATH = DATA_DIR / 'submissions.csv'
XLSX_PATH = DATA_DIR / 'submissions.xlsx'

app = Flask(__name__, static_folder='.', static_url_path='')

CREATE_SQL = """
CREATE TABLE IF NOT EXISTS visits (
  id INTEGER PRIMARY KEY CHECK (id = 1),
  total_views INTEGER NOT NULL DEFAULT 0,
  updated_at TEXT
);
CREATE TABLE IF NOT EXISTS submissions (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  created_at TEXT NOT NULL,
  name TEXT NOT NULL,
  phone TEXT NOT NULL,
  specialization TEXT,
  district TEXT,
  employment TEXT,
  consent INTEGER NOT NULL,
  ip_address TEXT,
  user_agent TEXT
);
INSERT OR IGNORE INTO visits (id, total_views, updated_at) VALUES (1, 0, NULL);
"""
HEADERS = ['created_at','name','phone','specialization','district','employment','consent','ip_address','user_agent']


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript(CREATE_SQL)
    conn.commit()
    conn.close()


def append_csv(row):
    write_header = not CSV_PATH.exists()
    with CSV_PATH.open('a', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        if write_header:
            writer.writeheader()
        writer.writerow(row)


def append_xlsx(row):
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Заявки'
        ws.append(HEADERS)
    ws.append([row[h] for h in HEADERS])
    wb.save(XLSX_PATH)


@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'povar.html')


@app.post('/api/visit')
def visit():
    conn = get_db()
    now = datetime.utcnow().isoformat()
    conn.execute('UPDATE visits SET total_views = total_views + 1, updated_at = ? WHERE id = 1', (now,))
    conn.commit()
    total = conn.execute('SELECT total_views FROM visits WHERE id = 1').fetchone()['total_views']
    conn.close()
    return jsonify({'ok': True, 'total_views': total})


@app.post('/api/submit')
def submit():
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    phone = (data.get('phone') or '').strip()
    specialization = (data.get('specialization') or '').strip()
    district = (data.get('district') or '').strip()
    employment = (data.get('employment') or '').strip()
    consent = bool(data.get('consent'))
    if not name or not phone:
        return jsonify({'ok': False, 'error': 'Имя и телефон обязательны'}), 400
    if not consent:
        return jsonify({'ok': False, 'error': 'Нужно согласие на обработку персональных данных'}), 400
    created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    row = {
        'created_at': created_at,
        'name': name,
        'phone': phone,
        'specialization': specialization,
        'district': district,
        'employment': employment,
        'consent': 'Да',
        'ip_address': request.headers.get('CF-Connecting-IP') or request.headers.get('X-Forwarded-For', request.remote_addr),
        'user_agent': request.headers.get('User-Agent', '')[:500],
    }
    conn = get_db()
    conn.execute(
        'INSERT INTO submissions (created_at, name, phone, specialization, district, employment, consent, ip_address, user_agent) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
        (created_at, name, phone, specialization, district, employment, 1, row['ip_address'], row['user_agent'])
    )
    conn.commit()
    conn.close()
    append_csv(row)
    append_xlsx(row)
    return jsonify({'ok': True})


@app.get('/api/stats')
def stats():
    conn = get_db()
    total_views = conn.execute('SELECT total_views FROM visits WHERE id = 1').fetchone()['total_views']
    total_forms = conn.execute('SELECT COUNT(*) AS cnt FROM submissions').fetchone()['cnt']
    latest = [dict(r) for r in conn.execute('SELECT created_at, name, phone, specialization, district, employment FROM submissions ORDER BY id DESC LIMIT 20').fetchall()]
    conn.close()
    return jsonify({'ok': True, 'total_views': total_views, 'total_forms': total_forms, 'latest': latest})


@app.get('/downloads/<path:filename>')
def downloads(filename):
    return send_from_directory(DATA_DIR, filename, as_attachment=True)


if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=8000, debug=False)
